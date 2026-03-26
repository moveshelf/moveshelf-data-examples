import os, json, copy
from moveshelf_api.api import MoveshelfApi
from typing import Any
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

## ============================================================================
## PREREQUISITES - Expected to be already defined in the script calling these functions
## ============================================================================
CONFIG_FOLDER = ""  # Folder with config and API key defined, typically the parent folder of this script

## Setup the API
# Load config
personal_config = os.path.join(CONFIG_FOLDER, "mvshlf-config.json")
if not os.path.isfile(personal_config):
    raise FileNotFoundError(
        f"Configuration file '{personal_config}' is missing.\n"
        "Ensure the file exists with the correct name and path."
    )

with open(personal_config, "r") as config_file:
    data = json.load(config_file)

custom_timeout = 120 # timeout (integer) for HTTP calls. Defaults to 120 if undefined
api = MoveshelfApi(
    api_key_file=os.path.join(CONFIG_FOLDER , data["apiKeyFileName"]),
    api_url=data["apiUrl"],
    timeout=custom_timeout,
)

## ============================================================================
## CONFIGURATION VARIABLES - Edit these before running
## ============================================================================
OUTPUT_FILENAME = "Metadata_Export.xlsx"

# Specify which fields to export (leave empty for all)
SUBJECT_METADATA_FIELDS: list[str] = ['ehr-id', 'subject-diagnosis']
SESSION_METADATA_FIELDS: list[str] = ['vicon-height', 'vicon-weight', 'interview-assistive-device']

# Column header format configuration
# When True: Use metadata field IDs as column headers
# (e.g., 'sessioninfo-comments', 'vicon-leg-length-right')
# When False: Use descriptive labels with tab context
# (e.g., 'Session info: Comments', 'Physical exam 1: Leg length (mm) - right')
USE_METADATA_ID_AS_COLUMN_HEADER: bool = True


## ============================================================================
## CONSTANTS
## ============================================================================

# Context labels for left/right fields
CONTEXT_LABELS = ["left", "right"]

# Value to use for missing data
DEFAULT_CELL_VALUE = ""

# Excel formatting constants
MAX_COLUMN_WIDTH = 50
COLUMN_PADDING = 5
HEADER_COLOR = '4472C4'
HEADER_FONT_COLOR = 'FFFFFF'
MAX_ROWS_FOR_WIDTH_CALC = 100


## ============================================================================
## MAIN EXPORT FUNCTION
## ============================================================================

def export_metadata_to_excel(
    project_id: str,
    output_filename: str = OUTPUT_FILENAME,
    subjects_list: list[dict[str, Any]] | None = None,
    sessions_list: list[dict[str, Any]] | None = None,
    subject_fields: list[str] | None = None,
    session_fields: list[str] | None = None,
    use_metadata_id_as_column_header: bool = USE_METADATA_ID_AS_COLUMN_HEADER
) -> None:
    """
    Export subject and session metadata to a formatted Excel file.
    
    This function retrieves metadata templates, processes subject and session data,
    and generates a well-formatted Excel file with descriptive column names.
    
    Can accept either:
    - A list of subjects (each with embedded sessions), OR
    - A list of sessions (which will be grouped by subject automatically)
    
    Args:
        project_id: ID of the Moveshelf project
        output_filename: Name of the output Excel file  
        subjects_list: Optional list of subject dictionaries with 'sessions' key
        sessions_list: Optional list of session dictionaries with 'patient' key
        subject_fields: List of subject metadata field keys to export. If None, uses config or all fields
        session_fields: List of session metadata field keys to export. If None, uses config or all fields
        use_metadata_id_as_column_header: If True, use metadata keys
                                          (e.g., 'sessioninfo-comments', 'vicon-leg-length-right')
                                          If False, use descriptive labels
                                          (e.g., 'Session info: Comments', 'Physical exam 1: Leg length (mm) - right')
        
    Raises:
        ValueError: If both subjects_list and sessions_list are provided or both are None
    """
    # Validate input
    if subjects_list and sessions_list:
        raise ValueError("Provide either subjects_list or sessions_list, not both")
    
    if not subjects_list and not sessions_list:
        raise ValueError("Must provide either subjects_list or sessions_list")
    
    # Use provided fields or fall back to config
    subject_fields = subject_fields or SUBJECT_METADATA_FIELDS
    session_fields = session_fields or SESSION_METADATA_FIELDS
    
    # Retrieve and process metadata template
    print(f"Retrieving metadata template for project {project_id}")
    metadata_template = api.getProjectTemplate(project_id)
    
    if not metadata_template:
        raise ValueError(f"Failed to retrieve metadata template for project {project_id}")
    
    subject_field_templates, session_field_templates = extract_metadata_fields(metadata_template)
    
    # Determine which fields to export (all if none specified)
    subject_fields_to_export = subject_fields if subject_fields else list(subject_field_templates.keys())
    session_fields_to_export = session_fields if session_fields else list(session_field_templates.keys())
    
    # Initialize data structure
    data_columns = define_column_names(
        subject_fields_to_export,
        subject_field_templates,
        session_fields_to_export,
        session_field_templates,
        use_metadata_id_as_column_header
    )
    
    # Prepare subjects list based on input type
    if sessions_list:
        print(f"Converting {len(sessions_list)} sessions to subject-grouped format")
        subjects_list = group_sessions_by_patient(sessions_list)
        print(f"Grouped into {len(subjects_list)} subjects")
    else:
        print(f"Processing {len(subjects_list)} subjects directly")
    
    # Process each subject and their sessions
    
    for subject in subjects_list:
        subject_id = subject.get('id', 'unknown')
        
        # Get subject metadata
        subject_metadata = ensure_subject_metadata(subject)
        
        # Get sessions for this subject
        subject_sessions = ensure_subject_sessions(subject)
        
        if not subject_sessions:
            print(f"Subject {subject_id} has no sessions, skipping")
            continue
        
        # Process each session
        for session in subject_sessions:
            session_metadata = ensure_session_metadata(session)
            
            # Append data for this row (one row per session)
            append_subject_metadata_to_row(
                data_columns,
                subject_metadata,
                subject_fields_to_export,
                subject_field_templates,
                use_metadata_id_as_column_header
            )
            append_session_date_and_label(
                data_columns,
                session
            )
            append_session_metadata_to_row(
                data_columns,
                session_metadata,
                session_fields_to_export,
                session_field_templates,
                use_metadata_id_as_column_header
            )
    
    # Create DataFrame and export to Excel
    print("Creating DataFrame and exporting to Excel")
    df = pd.DataFrame(data_columns)
    save_to_excel(df, output_filename, sheet_name='Metadata')
    
    print(f"Export completed: {len(df)} rows exported to {output_filename}")

## ============================================================================
## HELPER FUNCTIONS
## ============================================================================

def generate_column_name(field_key: str, field_info: dict[str, Any], use_metadata_id: bool = False) -> str:
    """
    Generate a unique, descriptive column name for a metadata field.
    
    Handles special cases like:
    - Score fields with generic labels (e.g., 'Score' -> 'Patient Goal 1 Score')
    - Interview fields with current/previous prefixes
    - Questionnaire fields with patient/parent response prefixes
    
    Args:
        field_key: The field identifier (e.g., 'interview-previous-patient-goal-1-score')
        field_info: Metadata dictionary containing 'tab', 'label', and other field properties
        use_metadata_id: If True, use metadata key directly instead of generating descriptive label
        
    Returns:
        str: Formatted column name (e.g., 'Interview: Previous Patient Goal 1 Score' or
        'interview-previous-patient-goal-1-score')
        
    Example:
        >>> generate_column_name('interview-previous-patient-goal-1-score', {'tab': 'Interview', 'label': 'Score'})
        'Interview: Previous Patient Goal 1 Score'
        >>> generate_column_name('vicon-height', {'tab': 'Physical Exam 1', 'label': 'Height (cm)'},
        use_metadata_id=True)
        'vicon-height'
    """
    if use_metadata_id:
        # Return the metadata key directly
        return field_key
    
    tab = field_info.get('tab', '')
    label = field_info.get('label', '')
    
    # Handle score fields with generic labels by extracting context from field key
    if field_key.endswith('-score') and label == 'Score':
        label = _generate_score_label(field_key)
    
    # Apply prefix based on field type
    prefix = _get_field_prefix(field_key)
    full_label = f"{prefix} {label}" if prefix else label
    
    # Combine with tab name
    return f"{tab}: {full_label}" if tab else full_label


def _generate_score_label(field_key: str) -> str:
    """
    Extract descriptive label for score fields from the field key.
    
    Args:
        field_key: Field key ending with '-score'
        
    Returns:
        str: Descriptive score label (e.g., 'Patient Goal 1 Score')
    """
    # Remove '-score' suffix and split into parts
    parts = field_key.rsplit('-score', 1)[0].split('-')
    
    # Look for goal patterns
    if 'goal' in parts:
        goal_idx = parts.index('goal')
        goal_number = parts[goal_idx + 1] if goal_idx + 1 < len(parts) else ''
        
        if 'patient' in parts:
            return f"Patient Goal {goal_number} Score"
        elif 'family' in parts:
            return f"Family Goal {goal_number} Score"
    
    # Default fallback
    return 'Score'


def _get_field_prefix(field_key: str) -> str:
    """
    Determine the appropriate prefix for a field based on its key.
    
    Args:
        field_key: The field identifier
        
    Returns:
        str: Prefix to add to the label (empty string if no prefix needed)
    """
    prefix_map = {
        'interview-current-': 'Current',
        'interview-previous-': 'Previous',
        'question-patient-': 'Patient Response',
        'question-parent-': 'Parent Response'
    }
    
    for prefix, label in prefix_map.items():
        if field_key.startswith(prefix):
            return label
    
    return ''

## ============================================================================
## METADATA TEMPLATE PROCESSING
## ============================================================================

def extract_metadata_fields(metadata_template: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    """
    Extract and organize subject and session metadata fields from project template.
    
    Args:
        metadata_template: Project metadata template from API
        
    Returns:
        Tuple of (subject_fields, session_fields) dictionaries
    """
    subject_fields = {}
    session_fields = {}
    
    # Extract subject metadata fields
    for field_key, field_value in metadata_template.get('template', {}).items():
        if field_value.get('type') != 'title':
            subject_fields[field_key] = field_value

    for tab in metadata_template.get('subjectMetadataTabs', []):
        # To extract 'interventions' and 'Research projects'
        tab_name = tab.get('name', '')
        subject_fields[tab_name] = {'label': tab_name}
    
    # Extract session metadata fields from tabs
    for tab in metadata_template.get('sessionMetadataTabs', []):
        tab_name = tab.get('name', '')
        for field_key, field_value in tab.get('template', {}).items():
            if field_value.get('type') != 'title':
                field_value['tab'] = tab_name
                session_fields[field_key] = field_value
    
    return subject_fields, session_fields


def define_column_names(
    subject_fields_to_export: list[str],
    subject_fields: dict[str, Any],
    session_fields_to_export: list[str],
    session_fields: dict[str, Any],
    use_metadata_id: bool = False
) -> dict[str, list]:
    """
    Create the initial data structure with column names as keys.
    
    Generates descriptive, unique column names for all fields to be exported,
    including special handling for context fields (Left/Right).
    
    Args:
        subject_fields_to_export: List of subject field keys to include
        subject_fields: Dictionary of subject field metadata
        session_fields_to_export: List of session field keys to include
        session_fields: Dictionary of session field metadata
        use_metadata_id: If True, use metadata keys as column names
        
    Returns:
        Dictionary with column names as keys and empty lists as values
    """
    data_columns: dict[str, list] = {}

    # Add subject columns
    for field_key in subject_fields_to_export:
        if field_key not in subject_fields:
            print(f"Subject field '{field_key}' not found in template, skipping")
            continue
        
        if use_metadata_id:
            column_name = field_key
        else:
            column_name = subject_fields[field_key].get('label', '')
        data_columns[column_name] = []

    # Add columns for session date and session label
    data_columns['Session Date'] = []
    data_columns['Session Label'] = []
    
    # Add session metadata columns
    for field_key in session_fields_to_export:
        if field_key not in session_fields:
            print(f"Session field '{field_key}' not found in template, skipping")
            continue
        
        field_info = session_fields[field_key]
        
        if field_info.get("hasContext", False):
            # Create separate columns for each context (Left/Right)
            _add_columns_with_context(data_columns, field_key, field_info, use_metadata_id)
        else:
            # Create single column with generated name
            column_name = generate_column_name(field_key, field_info, use_metadata_id)
            data_columns[column_name] = []

    return data_columns


def _add_columns_with_context(
    data_columns: dict[str, list],
    field_key: str,
    field_info: dict[str, Any],
    use_metadata_id: bool = False
) -> None:
    """
    Add columns for context-specific fields (left/right).
    
    Args:
        data_columns: Dictionary to add columns to
        field_key: The metadata field key
        field_info: Field metadata containing 'tab' and 'label'
        use_metadata_id: If True, use format 'field-key-left', else 'Tab: Label - left'
    """
    if use_metadata_id:
        # Format: metadata-key-left, metadata-key-right
        for context in CONTEXT_LABELS:
            column_name = f"{field_key}-{context}"
            data_columns[column_name] = []
    else:
        # Format: Tab: Label - left, Tab: Label - right
        tab = field_info.get('tab', '')
        label = field_info.get('label', '')
        
        for context in CONTEXT_LABELS:
            column_name = f"{tab}: {label} - {context}" if tab else f"{label} - {context}"
            data_columns[column_name] = []

## ============================================================================
## DATA POPULATION FUNCTIONS
## ============================================================================

def get_case_insensitive(dictionary: dict[str, Any], key: str, default: Any = None) -> Any:
    """
    Get a value from dictionary using case-insensitive key matching.
    
    First tries exact match, then falls back to case-insensitive search.
    
    Args:
        dictionary: Dictionary to search
        key: Key to look for
        default: Default value if key not found
        
    Returns:
        Value from dictionary or default
    """
    # Try exact match first (fast path)
    if key in dictionary:
        return dictionary[key]
    
    # Fall back to case-insensitive search
    key_lower = key.lower()
    for dict_key, value in dictionary.items():
        if dict_key.lower() == key_lower:
            return value
    
    return default


def append_subject_metadata_to_row(
    data_columns: dict[str, list],
    subject_metadata: dict[str, Any],
    subject_fields_to_export: list[str],
    metadata_template: dict[str, Any],
    use_metadata_id: bool = False
) -> None:
    """
    Append subject metadata values to the data columns.
    
    Args:
        data_columns: Dictionary with column names and value lists
        subject_metadata: Subject metadata dictionary
        subject_fields_to_export: List of field keys to export
        metadata_template: Field metadata template
        use_metadata_id: If True, use metadata key as column name
    """
    for field_key in subject_fields_to_export:
        if use_metadata_id:
            column_name = field_key
        else:
            column_name = metadata_template.get(field_key, {}).get('label', '')
        
        if not column_name or column_name not in data_columns:
            continue  # Skip invalid or filtered fields
        
        # Get value with case-insensitive key matching
        value = get_case_insensitive(subject_metadata, field_key, DEFAULT_CELL_VALUE)
        data_columns[column_name].append(value)


def append_session_metadata_to_row(
    data_columns: dict[str, list],
    session_metadata: dict[str, Any],
    session_fields_to_export: list[str],
    metadata_template: dict[str, Any],
    use_metadata_id: bool = False
) -> None:
    """
    Append session metadata values to the data columns.
    
    Handles both context fields (Left/Right) and regular fields with
    intelligent column name matching.
    
    Args:
        data_columns: Dictionary with column names and value lists
        session_metadata: Session metadata dictionary
        session_fields_to_export: List of field keys to export
        metadata_template: Field metadata template
        use_metadata_id: If True, use metadata key as column name
    """
    for field_key in session_fields_to_export:
        field_info = metadata_template.get(field_key, {})
        field_value = get_case_insensitive(session_metadata, field_key, None)
        
        if field_info.get("hasContext", False):
            _append_field_with_context(data_columns, field_value, field_key, field_info, use_metadata_id)
        else:
            _append_regular_field(data_columns, field_value, field_key, field_info, use_metadata_id)


def format_date_to_yyyy_mm_dd(date_string: str) -> str:
    """
    Convert ISO format date to yyyy-mm-dd format.
    
    Args:
        date_string: Date string in ISO 8601 format 
        (e.g., '2025-07-09T00:00:00+00:00')
        
    Returns:
        Date string in yyyy-mm-dd format (e.g., '2025-07-09'),
        or original string if parsing fails
    """
    if not date_string or date_string == DEFAULT_CELL_VALUE:
        return date_string
    
    try:
        # Parse ISO format and convert to yyyy-mm-dd
        dt = datetime.fromisoformat(date_string.replace('Z', '+00:00'))
        return dt.strftime('%Y-%m-%d')
    except (ValueError, AttributeError):
        # If parsing fails, return original string
        return date_string


def append_session_date_and_label(
    data_columns: dict[str, list],
    session: dict[str, Any]
) -> None:
    """
    Append session date and label to the data columns.
    
    Args:
        data_columns: Dictionary with column names and value lists
        session: Session dictionary containing 'Session Date' and 'Session Label' keys
    """
    session_date = session.get('date', DEFAULT_CELL_VALUE)
    # Convert date to yyyy-mm-dd format
    session_date = format_date_to_yyyy_mm_dd(session_date)
    
    if 'Session Date' in data_columns:
        data_columns['Session Date'].append(session_date)

    # Extract session label from session metadata
    metadata_str = session.get("metadata", None)
    
    # Fetch metadata if missing
    if not metadata_str:
        session_id = session.get('id')
        print(f"Fetching metadata for session {session_id}")
        session_details = api.getSessionById(session_id)
        
        if session_details and 'metadata' in session_details:
            metadata_str = session_details['metadata']
        else:
            print(f"Failed to retrieve metadata for session {session_id}")
            metadata_str = "{}"
    
    # Parse JSON and extract session label
    try:
        metadata = json.loads(metadata_str)
        session_label = metadata.get('label', DEFAULT_CELL_VALUE)
    except json.JSONDecodeError:
        print(f"Invalid JSON metadata for session {session.get('id', 'unknown')}")
        session_label = DEFAULT_CELL_VALUE
    
    if 'Session Label' in data_columns:
        data_columns['Session Label'].append(session_label)


def _append_field_with_context(
    data_columns: dict[str, list],
    field_value: list[dict[str, Any]] | None,
    field_key: str,
    field_info: dict[str, Any],
    use_metadata_id: bool = False
) -> None:
    """
    Append values for a context field (left/right) to appropriate columns.
    
    Args:
        data_columns: Dictionary with column names and value lists
        field_value: List of context dictionaries with 'context' and 'value' keys
        field_key: The metadata field key
        field_info: Field metadata containing 'tab' and 'label'
        use_metadata_id: If True, use format 'field-key-left', else 'Tab: Label - left'
    """
    # Create a dictionary for fast context lookup
    context_values = {}
    if field_value:
        context_values = {
            item.get('context', '').lower(): item.get('value', DEFAULT_CELL_VALUE)
            for item in field_value
        }
    
    # Append value for each context
    for context in CONTEXT_LABELS:
        if use_metadata_id:
            column_name = f"{field_key}-{context}"
        else:
            tab = field_info.get('tab', '')
            label = field_info.get('label', '')
            column_name = f"{tab}: {label} - {context}" if tab else f"{label} - {context}"
        
        if column_name in data_columns:
            value = context_values.get(context.lower(), DEFAULT_CELL_VALUE)
            data_columns[column_name].append(value)


def _append_regular_field(
    data_columns: dict[str, list],
    field_value: Any,
    field_key: str,
    field_info: dict[str, Any],
    use_metadata_id: bool = False
) -> None:
    """
    Append value for a regular (non-context) field to the appropriate column.
    
    Args:
        data_columns: Dictionary with column names and value lists
        field_value: The field value to append
        field_key: The field identifier
        field_info: Field metadata
        use_metadata_id: If True, use metadata key as column name
    """
    column_name = generate_column_name(field_key, field_info, use_metadata_id)
    
    if column_name in data_columns:
        value = field_value if field_value is not None else DEFAULT_CELL_VALUE
        data_columns[column_name].append(value)

## ============================================================================
## DATA TRANSFORMATION FUNCTIONS
## ============================================================================

def group_sessions_by_patient(sessions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Group sessions by patient/subject, converting a flat session list into a hierarchical structure.
    
    Takes a list of session dictionaries (each containing patient info) and reorganizes
    them into a list of patient dictionaries, each with a 'sessions' list.
    
    Args:
        sessions: List of session dictionaries with 'patient' key
        
    Returns:
        List of patient dictionaries with 'sessions' key containing their sessions
        
    Example:
        Input: [{'id': 's1', 'patient': {'id': 'p1', 'name': 'John'}}, 
                {'id': 's2', 'patient': {'id': 'p1', 'name': 'John'}}]
        Output: [{'id': 'p1', 'name': 'John', 'sessions': [{'id': 's1'}, {'id': 's2'}]}]
    """
    patients_dict = {}

    for session in sessions:
        patient_info = session.get("patient", {})
        patient_id = patient_info.get("id")
        
        if not patient_id:
            print(f"Session {session.get('id', 'unknown')} has no patient ID, skipping")
            continue

        # Initialize patient entry if not seen before
        if patient_id not in patients_dict:
            patients_dict[patient_id] = copy.deepcopy(patient_info)
            patients_dict[patient_id]["sessions"] = []

        # Add session without duplicate patient info
        session_copy = {k: v for k, v in session.items() if k != "patient"}
        patients_dict[patient_id]["sessions"].append(session_copy)

    return list(patients_dict.values())


def ensure_subject_metadata(subject: dict[str, Any]) -> dict[str, Any]:
    """
    Ensure subject has metadata, fetching from API if necessary.
    
    Args:
        subject: Subject dictionary
        
    Returns:
        Parsed metadata dictionary (empty dict if not available)
    """
    metadata_str = subject.get("metadata", None)
    
    # Fetch metadata if missing
    if not metadata_str:
        subject_id = subject.get('id')
        print(f"Fetching metadata for subject {subject_id}")
        subject_details = api.getSubjectDetails(subject_id)
        
        if subject_details and 'metadata' in subject_details:
            metadata_str = subject_details['metadata']
        else:
            print(f"Failed to retrieve metadata for subject {subject_id}")
            return {}
    
    # Parse JSON metadata
    try:
        return json.loads(metadata_str)
    except json.JSONDecodeError:
        print(f"Invalid JSON metadata for subject {subject.get('id', 'unknown')}")
        return {}


def ensure_session_metadata(session: dict[str, Any]) -> dict[str, Any]:
    """
    Ensure session has metadata, fetching from API if necessary.
    
    Args:
        session: Session dictionary
        
    Returns:
        Parsed metadata dictionary (empty dict if not available)
    """
    metadata_str = session.get("metadata", None)
    
    # Fetch metadata if missing
    if not metadata_str:
        session_id = session.get('id')
        print(f"Fetching metadata for session {session_id}")
        session_details = api.getSessionById(session_id)
        
        if session_details and 'metadata' in session_details:
            metadata_str = session_details['metadata']
        else:
            print(f"Failed to retrieve metadata for session {session_id}")
            return {}
    
    # Parse JSON and extract nested metadata
    try:
        metadata = json.loads(metadata_str)
        return metadata.get('metadata', {})
    except json.JSONDecodeError:
        print(f"Invalid JSON metadata for session {session.get('id', 'unknown')}")
        return {}


def ensure_subject_sessions(subject: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Ensure subject has sessions list, fetching from API if necessary.
    
    Args:
        subject: Subject dictionary
        
    Returns:
        List of session dictionaries
    """
    sessions = subject.get('sessions', [])
    
    if not sessions:
        subject_id = subject.get('id')
        print(f"Fetching sessions for subject {subject_id}")
        subject_data = api.getSubjectData(subject_id)
        
        if subject_data and 'sessions' in subject_data:
            sessions = subject_data['sessions']
        else:
            print(f"No sessions found for subject {subject_id}")
    
    return sessions


## ============================================================================
## EXCEL EXPORT AND FORMATTING
## ============================================================================

def save_to_excel(df: pd.DataFrame, filename: str, sheet_name: str = 'Data') -> None:
    """
    Save DataFrame to Excel with professional formatting.
    
    Applies:
    - Colored header row with white text
    - Auto-sized columns (with max width limit)
    - Frozen top row for scrolling
    - Table format with filtering enabled
    
    Args:
        df: DataFrame to export
        filename: Output file path
        sheet_name: Name of the Excel sheet (default: 'Data')
    """
    if df.empty:
        print(f"DataFrame is empty, no data to export to '{filename}'")
        return
    
    # Write DataFrame to Excel
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Apply formatting
    _apply_excel_formatting(filename, sheet_name, df)
    
    print(f"Excel file saved successfully: {filename}")


def _apply_excel_formatting(filename: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Apply formatting to an existing Excel file.
    
    Args:
        filename: Path to Excel file
        sheet_name: Name of sheet to format
        df: DataFrame used to determine dimensions
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    
    # Format header row
    _format_header_row(ws)
    
    # Auto-size columns
    _auto_size_columns(ws, df)
    
    # Add filter table
    _add_filter_table(ws, df, sheet_name)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filename)


def _format_header_row(worksheet) -> None:
    """
    Apply formatting to the header row (row 1).
    
    Args:
        worksheet: Openpyxl worksheet object
    """
    header_fill = PatternFill(
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
        fill_type='solid'
    )
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _auto_size_columns(worksheet, df: pd.DataFrame) -> None:
    """
    Automatically size columns based on content width.
    
    Args:
        worksheet: Openpyxl worksheet object
        df: DataFrame used to calculate column widths
    """
    for col_num, col_header in enumerate(df.columns, 1):
        # Start with header width
        max_length = len(str(col_header))
        
        # Check data rows (limit check to first 100 rows for performance)
        max_row = min(worksheet.max_row, MAX_ROWS_FOR_WIDTH_CALC + 1)
        for row_num in range(2, max_row + 1):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width with padding and max limit
        col_letter = get_column_letter(col_num)
        adjusted_width = min(max_length + COLUMN_PADDING, MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def _add_filter_table(worksheet, df: pd.DataFrame, sheet_name: str) -> None:
    """
    Add Excel table with filters.
    
    Args:
        worksheet: Openpyxl worksheet object
        df: DataFrame used to determine table range
        sheet_name: Sheet name (used for table name)
    """
    if df.empty:
        return
    
    # Define table range (A1 to last column/row)
    last_col = get_column_letter(df.shape[1])
    last_row = df.shape[0] + 1
    table_ref = f'A1:{last_col}{last_row}'
    
    # Create and add table
    table_name = sheet_name.replace(' ', '_').replace('-', '_')
    table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_ref)
    worksheet.add_table(table)