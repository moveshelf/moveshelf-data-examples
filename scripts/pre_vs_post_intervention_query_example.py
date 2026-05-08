from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
import copy
import os
import re
import sys
import json
import logging
from typing import Any
from collections import Counter
from matplotlib.dates import relativedelta
import numpy as np
import pandas as pd
import requests
parent_folder = os.path.dirname(os.path.dirname(__file__))
sys.path.append(parent_folder)
from api.api_additions import MoveshelfApiCustomized
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


# Use a requests.Session for connection pooling
MAX_WORKERS = 5 # Number of threads for parallel processing.
POOL_MAXSIZE = MAX_WORKERS + 2  # Set pool maxsize slightly higher than max workers to avoid connection issues
requests_session = requests.Session()
adapter = requests.adapters.HTTPAdapter(pool_maxsize=POOL_MAXSIZE)
requests_session.mount('https://', adapter)

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

## ============================================================================
## README: Pre vs Post Intervention Query Example
## ============================================================================
#
# This script demonstrates how to perform a pre vs post intervention query on Moveshelf data,
# combining server-side filtering with local filtering to efficiently extract
# data from subjects who underwent interventions.
#
## PURPOSE:
# The script queries one or more Moveshelf projects for subjects with at least two sessions
# (pre and post intervention), applies inclusion/exclusion criteria at multiple levels
# (subject, session, and clip), and exports selected parameters to Excel files.
#
## OUTPUT:
# Two Excel files are generated:
# 1. DATA_SPREADSHEET_FILENAME: Contains subject demographics, session metadata, and kinematic
#    data (angles, lengths, velocities, gait parameters) for pre and post sessions.
# 2. INTERVENTION_DATA_SPREADSHEET_FILENAME: Contains detailed intervention/procedure information
#    for each subject.
#
## FILTERING LOGIC (Three-Layer approach):
#
# The script uses a combination of server-side and local filtering to optimize performance:
#
# 1. SERVER-SIDE FILTERING (Initial Query):
#    - Applied via `subject_metadata_filters` and `session_filters` parameters
#    - Filters subjects by subject metadata (e.g., diagnosis) and sessions by date range and count
#    - Reduces the dataset size before downloading, minimizing API calls and data transfer
#    - Returns subjects WITHOUT clip-level data for efficiency
#
# 2. LOCAL FILTERING - LAYER 1 (Subject Level):
#    - Applied using `INCLUSION_CRITERIA` with level="subject"
#    - Filters subjects based on subject metadata (e.g., diagnosis, date of birth)
#    - Can be skipped if already covered by server-side filtering
#
# 3. LOCAL FILTERING - LAYER 2 (Session Level and interventions):
#    - Applied using `INCLUSION_CRITERIA` with level="session" and `INTERVENTION_CRITERIA`
#    - Identifies valid consecutive session pairs (pre/post) based on:
#      a. Time constraints: sessions must be MIN_MONTHS_BETWEEN_SESSIONS apart
#      b. Intervention criteria: filters by procedures between sessions using INTERVENTION_CRITERIA
#      c. Session metadata: validates session-specific metadata (e.g., popliteal angle)
#    - Assigns subjects to intervention groups (e.g., "MHL", "MLHL", "Control") based on
#      procedures performed between the pre and post sessions. Control group includes subjects without qualifying procedures 
#      in between sessions.
#    - Exclusion criteria can remove subjects with unwanted procedures
#
# 4. LOCAL FILTERING - LAYER 3 (Clip Level):
#    - Applied using `INCLUSION_CRITERIA` with level="clip"
#    - Downloads clip-level JSON data only for subjects passing previous filters
#    - Evaluates kinematic data (angles, lengths, velocities) from specific conditions
#    - Most resource-intensive step, so performed last to minimize unnecessary processing
#
## CONTEXT-BASED INCLUSION CRITERIA AND LATERALITY:
#
# Many parameters in gait analysis are context-specific (left/right side). When using
# context-based inclusion criteria (has_context=True), the script applies special logic
# to ensure consistency with the subject's impairment laterality:
#
# - If subject metadata contains "subject-diagnosis-laterality" = "Left" or "Right":
#   The script ONLY evaluates criteria for that specific side. This ensures that filtering
#   is based on the impaired/affected limb, not the unaffected limb.
#
# - If laterality is not specified, "Both", or empty:
#   The script evaluates BOTH sides and determines which side(s) pass all criteria:
#   * If both sides pass all context-based criteria → final_side = "right" (default to right if both sides are valid)
#   * If only the right side passes all criteria → final_side = "right"
#   * If only the left side passes all criteria → final_side = "left"
#   * If different sides pass different criteria (mixed) → subject is EXCLUDED
#
# This ensures that:
# 1. All context-based criteria are evaluated on the same side (left, right, or both)
# 2. For subjects with unilateral impairments, filtering focuses on the affected limb
# 3. Results are consistent and clinically meaningful
#
# Example: If a criterion requires "Knee angle at initial contact > 5°" with has_context=True,
# and the subject has Left-sided CP:
# - Only the LEFT knee angle is checked
# - The right knee angle is ignored
# - The subject passes if the left knee meets the criterion
#
## SESSION PAIR LOGIC:
#
# The script processes all consecutive session pairs for each subject:
# - For a subject with 3 sessions [S1, S2, S3], it evaluates pairs: (S1,S2) and (S2,S3)
# - Each pair is validated independently against all criteria
# - Multiple valid pairs from the same subject can be included (treated as separate rows)
#
## CUSTOMIZATION:
#
# To adapt this script for other studies, modify the configuration variables below:
# - PARAMS_TO_EXPORT: Define which parameters to extract and export
# - INCLUSION_CRITERIA: Define filtering rules at subject/session/clip levels
# - INTERVENTION_CRITERIA: Define which interventions to include/exclude
# - CONDITION_TARGET_NAMES: Specify which trial conditions to analyze (or [] for all)
# - Date ranges, time windows, and other constraints
#
## CURRENT LIMITATIONS:
# - Single support average processing for kinematic data is not yet implemented


## ============================================================================
## Configuration variables - edit these before running the script. 
## Refer to comments for each variable for instructions on how to define them.
## ============================================================================

query_projects_in_parallel = False # If True, use multithreading to query multiple projects in parallel

PROJECT_NAMES = ['<organizationName/projectName1>', '<organizationName/projectName2>']  # List of project names to query

DEFAULT_CELL_VALUE = "" # If a field is not found, assign ""
CONTEXT_LABELS = ["Left", "Right"] # Labels used for context-specific parameters (e.g., left/right)

DATA_SPREADSHEET_FILENAME = 'Data Export - Pre vs post intervention study.xlsx'
INTERVENTION_DATA_SPREADSHEET_FILENAME = 'Intervention Data Export - Pre vs post intervention study.xlsx'

# Excel formatting constants
MAX_COLUMN_WIDTH = 50
COLUMN_PADDING = 5
HEADER_COLOR = '4472C4'
HEADER_FONT_COLOR = 'FFFFFF'
MAX_ROWS_FOR_WIDTH_CALC = 100

# Define parameters that should be exported to DATA_SPREADSHEET_FILENAME. 
# Define an empty list if no parameters should be exported. Each parameter is defined as a dictionary with the following keys:
# - column_name: name of the column in the output excel file
# - level: project, subject, session, clip
# - source: name of the source where the parameter is stored (e.g., name, subject_metadata, session_metadata, angles_normalized, lengths_velocities_normalized, gait_params)
# - label: name of the label (e.g. metadata key) in the source (not needed if source is "name" or "date")
# - processing: for clip level parameters, specify if the value should be taken at the start, average, or max of the signal. None if no processing is needed (e.g., for gait_params)
# - has_context: True if the parameter has context (e.g., left/right), False otherwise    
PARAMS_TO_EXPORT = [
    # Subject-level
    {"column_name": "Subject name", "level": "subject", "source": "name", "has_context": False},
    {"column_name": "Date of birth", "level": "subject", "source": "subject_metadata", "label": "subject-date-of-birth", "has_context": False},
    {"column_name": "Diagnosis", "level": "subject", "source": "subject_metadata", "label": "subject-diagnosis", "has_context": False},
    {"column_name": "Laterality", "level": "subject", "source": "subject_metadata", "label": "subject-diagnosis-laterality", "has_context": False},
    {"column_name": "Limb used for filtering", "level": "subject", "source": "passed_side", "has_context": False},
    {"column_name": "Group", "level": "subject", "source": "group", "has_context": False},
    # Session-level
    {"column_name": "Session date", "level": "session", "source": "date", "has_context": False},
    {"column_name": "Age at encounter", "level": "session", "source": "date", "processing": "age_at_encounter", "has_context": False},
    {"column_name": "Conditions collected", "level": "session", "source": "session_metadata", "label": "sessioninfo-conditions-collected", "has_context": False},
    {"column_name": "Data collected", "level": "session", "source": "session_metadata", "label": "sessioninfo-data-collected", "has_context": False},
    {"column_name": "Primary orthoses", "level": "session", "source": "session_metadata", "label": "interview-orthotics", "has_context": True},
    {"column_name": "Height (cm)", "level": "session", "source": "session_metadata", "label": "vicon-height", "has_context": False},
    {"column_name": "Weight (kg)", "level": "session", "source": "session_metadata", "label": "vicon-weight", "has_context": False},
    {"column_name": "Leg length (mm)", "level": "session", "source": "session_metadata", "label": "vicon-leg-length", "has_context": True},
    # Clip-level parameters (only for clips with conditions in CONDITION_TARGET_NAMES)
    {"column_name": "Knee Angle at Initial Contact", "level": "clip", "source": "angles_normalized", "label": "KneeFlexExt",  "processing":"start", "has_context": True},
    {"column_name": "Peak knee extension in gait cycle", "level": "clip", "source": "angles_normalized", "label": "KneeFlexExt",  "processing":"max", "has_context": True},
    {"column_name": "Peak Ankle Dorsiflexion in gait cycle", "level": "clip", "source": "angles_normalized", "label": "DorsiPlanFlex",  "processing":"max", "has_context": True},
    {"column_name": "Medial Hamstring length at Initial Contact", "level": "clip", "source": "lengths_velocities_normalized", "label": "MedHamstringLength",  "processing":"start", "has_context": True},
    {"column_name": "Peak Medial Hamstring velocity in gait cycle", "level": "clip", "source": "lengths_velocities_normalized", "label": "MedHamstringVelocity",  "processing":"max", "has_context": True},
    {"column_name": "Walking speed", "level": "clip", "source": "gait_params", "label": "Speed",  "processing": None,  "has_context": False},
]

# Define parameters that should be exported to INTERVENTION_DATA_SPREADSHEET_FILENAME.
# Define an empty list if no intervention parameters should be exported.
INTERVENTION_PARAMS_TO_EXPORT = [
    {"column_name": "Subject name", "level": "subject", "source": "name", "has_context": False},
    {"column_name": "Site", "level": "subject", "source": "intervention_metadata", "label": "site", "has_context": False},
    {"column_name": "Date", "level": "subject", "source": "intervention_metadata", "label": "date", "has_context": False},
    {"column_name": "Surgeon", "level": "subject", "source": "intervention_metadata", "label": "surgeon", "has_context": False},
    {"column_name": "Procedure side", "level": "subject", "source": "procedure_metadata", "label": "side", "has_context": False},
    {"column_name": "Procedure", "level": "subject", "source": "procedure_metadata", "label": "procedure", "has_context": False},
    {"column_name": "Procedure location", "level": "subject", "source": "procedure_metadata", "label": "location", "has_context": False},
    {"column_name": "Location modifier", "level": "subject", "source": "procedure_metadata", "label": "location-modifier", "has_context": False},
]

# Define inclusion criteria for filtering subjects/sessions/clips locally. Stored as a list of dictionaries with the following keys:
# - level: subject, session, clip
# - source: name of the source where the parameter is stored (e.g., subject_metadata, session_metadata, angles_normalized, lengths_velocities_normalized, gait_params)
# - label: name of the label in the source (not needed if source is "name" or "date")
# - value: list of values to compare against
# - operation: type of comparison (is, is_not, greater_than, less_than)
# - has_context: True if the parameter has context (e.g., left/right), False otherwise
# - pre_post: "pre" or "post" if the criterion should be applied to a specific session in a subject with multiple sessions (e.g., pre-operative session), None otherwise
INCLUSION_CRITERIA = [
    {"level": "subject", "source": "subject_metadata", "label": "subject-diagnosis", "value": ["Cerebral Palsy"], "operation": "is", "has_context": False}, # Diagnosis is Cerebral Palsy
    {"level": "session", "source": "session_metadata", "label": "supine-knee-single-pop-r1-angle", "value": ["35"], "operation": "greater_than", "has_context": True, "pre_post": "pre"}, # Popliteal angle greater than 35
    {"level": "clip", "source": "angles_normalized", "label": "KneeFlexExt", "value": ["5"], "processing": "start", "operation": "greater_than", "has_context": True, "pre_post": "pre"}, # Knee angle at initial contact greater than 5
]

# Define intervention criteria for filtering subjects based on their interventions. Stored as a dictionary with "inclusion" and "exclusion" keys.
# Each key contains a list of lists of dictionaries with the following keys:
# - label: name of the intervention/procedure metadata key
# - value: list of values to compare against
# - operation: type of comparison (is, is_not, greater_than, less_than)
# - group: name of the group the criterion belongs to
# - all criteria in a list must be met (AND), at least one list must be met (OR between lists)
INTERVENTION_CRITERIA = {
    "inclusion":[
        [
            {"label": "location-modifier", "value": ["Medial", ""], "operation": "is", "group": "Intervention A"},
            {"label": "location", "value": ["Hamstrings", "Semimembranosus", "Semitendinosus", "Gracilis"], "operation": "is", "group": "Intervention A"},
            {"label": "procedure", "value": ["Lengthening"], "operation": "is", "group": "Intervention A"}
        ],
        [
            {"label": "location-modifier", "value": ["Medial & Lateral", ""], "operation": "is", "group": "Intervention B"},
            {"label": "location", "value": ["Hamstrings", "Biceps Femoris"], "operation": "is", "group": "Intervention B"},
            {"label": "procedure", "value": ["Lengthening"], "operation": "is", "group": "Intervention B"}
        ],
    ],
    "exclusion":[
        [
            {"label": "location", "value": ["Gastrocnemius", "Gastrocnemius-Soleus", "Achilles Tendon"], "operation": "is", "group": "Exclusion"},
            {"label": "procedure", "value": ["Lengthening", "Recession"], "operation": "is", "group": "Exclusion"}
        ],
        [
            {"label": "location-modifier", "value": ["Distal"], "operation": "is", "group": "Exclusion"},
            {"label": "location", "value": ["Femur"], "operation": "is", "group": "Exclusion"},
            {"label": "procedure", "value": ["Osteotomy"], "operation": "is", "group": "Exclusion"},
            {"label": "procedure-modifier", "value": ["Extension"], "operation": "is", "group": "Exclusion"}
        ]
    ]
}

STARTING_DATE = None  # Specifies the earliest session date considered in query. String in "YYYY-MM-DD" format. Set to None if no starting date filter is needed.
ENDING_DATE = "2020-08-01"    # Specifies the latest session date considered in query. String in "YYYY-MM-DD" format. Set to None if no ending date filter is needed.
STARTING_DATE_DT = datetime.fromisoformat(STARTING_DATE) if STARTING_DATE else None
ENDING_DATE_DT = datetime.fromisoformat(ENDING_DATE) if ENDING_DATE else None

MIN_MONTHS_BETWEEN_SESSIONS = 9  # Minimum number of months between pre and post sessions to be included. Set to None if no minimum is required.
MAX_YEARS_AFTER_INTERVENTION = 5  # Maximum number of years after intervention for the post session to be included. Set to None if no maximum is required.

# Define subject metadata filters for querying subjects with server-side filtering. Set to None if no server-side filtering is needed at the subject level. Refer to https://moveshelf.com/docs/api#retrievefilteredsubjects to define the filters.
subject_metadata_filters = {
        "key": "subject-diagnosis",
        "operator": "EQ",
        "value": "Cerebral Palsy"
    }

# Define session filters for querying sessions with server-side filtering
session_filters = {
    "sessionDates": {
        "startDate": STARTING_DATE,
        "endDate": ENDING_DATE,
    },
    "numSessions": {
        "min": 2
    }
}

CONDITION_TARGET_NAMES = ["barefoot", "bf walk"] # Names of conditions (lowercase) to process at the clip level. Only clips/trials within these conditions will be included in the export. Set to an empty list [] to include clips from all conditions.
# Ensure CONDITION_TARGET_NAMES is always lowercase for case-insensitive comparison
CONDITION_TARGET_NAMES = [name.lower() for name in CONDITION_TARGET_NAMES]

PRE_POST_SUFFIXES = [" - Pre", " - Post"] # Suffixes to add to column names if pre and post sessions are considered
    
## ============================================================================
## HELPER FUNCTIONS
## ============================================================================

def download_with_session(url: str) -> dict | None:
    return download_json_file(url, session=requests_session)

def download_json_file(url: str, session: requests.Session | None = None) -> dict | None:
    """
    Downloads and parses a JSON file from a URL.
    
    Args:
        url: URL to download from
        session: Optional requests session for connection pooling
    
    Returns:
        Parsed JSON data as dictionary, or None if download/parsing fails
    """
    try:
        response = session.get(url) if session else requests.get(url)
        decoded_content = response.content.decode()
        return json.loads(decoded_content)
    except Exception as e:
        print(f"Failed to download or parse {url}: {e}")
        return None
    
def define_column_names(export_params: list[dict]) -> dict[str, list]:
    """
    Defines the column names for the output spreadsheet based on the export parameters.
    Returns a dictionary with column names as keys and empty lists as values.
    """
    data_columns: dict[str, list] = {}

    for param in export_params:
        col_base = param["column_name"]

        if param.get("level", "") in ["project", "subject"]:
            # no need to add prefix or suffix
            data_columns[col_base] = []
        else:
            for pre_post in PRE_POST_SUFFIXES:
                if param.get("has_context"):
                    for context in CONTEXT_LABELS:
                        data_columns[f"{context} {col_base}{pre_post}"] = []
                else:
                    data_columns[f"{col_base}{pre_post}"] = []

    return data_columns

def calculate_age_in_years_and_months(date_of_birth: str, encounter_date: str) -> str:
    """
    Calculate age between two dates and return formatted as "X years Y months".
    Args:
        date_of_birth: Date of birth string in "YYYY-MM-DD" format.
        encounter_date: Encounter date string in "YYYY-MM-DD" format.
    Returns:
        Formatted age string like "12 years 6 months" or DEFAULT_CELL_VALUE if calculation fails.
    """
    try:
        dob = datetime.strptime(date_of_birth, "%Y-%m-%d")
        encounter = datetime.strptime(encounter_date, "%Y-%m-%d")
        
        # Calculate years and months
        years = encounter.year - dob.year
        months = encounter.month - dob.month
        
        # Adjust if we haven't reached the birth day in the current month
        if encounter.day < dob.day:
            months -= 1
        
        # Adjust if months is negative
        if months < 0:
            years -= 1
            months += 12
        
        return f"{years} years {months} months"
    except (ValueError, AttributeError):
        return DEFAULT_CELL_VALUE

def group_inclusion_criteria_by_source(criteria: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """
    Groups inclusion criteria by their source.
    Returns a dictionary with source names as keys and lists of criteria as values.
    """
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for criterion in criteria:
        grouped[criterion["source"]].append(criterion)

    return dict(grouped)

def group_clip_params_by_source(params: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """
    Groups clip-level parameters by their source.
    Returns a dictionary with source names as keys and lists of parameters as values.
    """
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for param in params:
        if param.get("level") == "clip":
            grouped[param["source"]].append(param)

    return dict(grouped)

def get_clip_data_sources(params_to_export: list[dict[str, Any]] | None = None, inclusion_criteria: list[dict[str, Any]] | None = None) -> set[str]:
    """
    Extracts unique source names from clip-level parameters and inclusion criteria.
    Returns a set of source names (e.g., {'angles_normalized', 'lengths_velocities_normalized', 'gait_params'})
    """
    sources = set()
    
    # Extract from params_to_export
    if params_to_export is not None:
        for param in params_to_export:
            if param.get("level") == "clip":
                source = param.get("source")
                if source:  # Only add non-empty sources
                    sources.add(source)
    
    # Extract from inclusion_criteria
    if inclusion_criteria is not None:
        for criterion in inclusion_criteria:
            if criterion.get("level") == "clip":
                source = criterion.get("source")
                if source:  # Only add non-empty sources
                    sources.add(source)
    
    return sources


def filter_based_on_inclusion_criteria(data_to_validate, inclusion_criteria, source, i_session: int | None = None, affected_side: str | None = None):
    """
    Checks if the given data meets all the inclusion criteria for the specified source.
    
    This function handles both context-agnostic and context-specific (left/right) criteria.
    For context-specific criteria, it respects the affected_side parameter to ensure
    filtering is done on the correct limb.
    
    Args:
        data_to_validate: The data to be validated against the criteria.
        inclusion_criteria: List of inclusion criteria dictionaries.
        source: The source name to filter criteria.
        i_session: Index of the session (0 for pre, 1 for post) if applicable.
        affected_side: diagnosis-laterality. If it is specified as 'left' or 'right', criteria with context will only check that side. Otherwise, both sides are evaluated.
    Returns True if all criteria are met, False otherwise. If context-aware, returns (True, side), where side can be "left", "right", or "both" or (False, None).
    """
    sides_passed_counter = Counter() # counts how many criteria with context passed for each side. At the end, if both sides have counts equal to the number of context-aware criteria, then both sides passed.
    context_criteria_count = 0 # number of criteria with context evaluated

    # Get required clip data sources dynamically
    clip_data_sources = get_clip_data_sources(PARAMS_TO_EXPORT, INCLUSION_CRITERIA) 

    for criterion in inclusion_criteria:
        # Only process criteria matching the current source
        if criterion["source"] != source:
            continue

        # Handle pre/post session logic - skip criteria not applicable to current session
        if criterion.get("pre_post") is not None and i_session is not None:
            if criterion["pre_post"] == "pre" and i_session != 0:
                continue  # This criterion is for pre-session only, skip for post
            elif criterion["pre_post"] == "post" and i_session != 1:
                continue  # This criterion is for post-session only, skip for pre

        # Extract value from data based on source type
        if source in ["subject_metadata", "session_metadata"]:
            # For metadata sources, extract value by key/label
            key = criterion.get("label")
            if not key or key not in data_to_validate:
                return False  # Required key missing
            
            value = data_to_validate.get(key, None)
            
            # Handle special case of multiselect metadata
            if isinstance(value, dict) and value.get("multiselect", False):
                value = value.get("value", [])
            
            if value is None:
                return False  # Value is None, criterion cannot be evaluated

        elif source in clip_data_sources:
            # For clip data sources (angles, lengths, gait params)
            # Build list of expected channel names based on context
            if criterion.get("has_context", False):
                # Context-aware: look for LeftLabel and RightLabel
                keys = [f"{context}{criterion['label']}" for context in CONTEXT_LABELS]
            else:
                # Context-agnostic: just the label itself
                keys = [criterion["label"]]
            
            # Extract matching items with their context (left/right)
            value = [
                {
                    "value": item.get("value"),
                    "context": "left" if item.get("channel_name", "").lower().startswith("left") else "right"
                }
                for item in data_to_validate
                if item.get("channel_name") in keys
            ]
            
            if not value:
                return False  # No matching data found
        else:
            # Source not recognized, skip this criterion
            continue

        criterion_values = criterion["value"] if isinstance(criterion["value"], list) else [criterion["value"]]
        match criterion["operation"]:
            case "is":
                value_str = str(value).lower()
                if not any(value_str == str(cv).lower() for cv in criterion_values):
                    return False
            case "is_not":
                value_str = str(value).lower()
                if any(value_str == str(cv).lower() for cv in criterion_values):
                    return False
            case "contains":
                value_lower = [str(item).lower() for item in value]
                if not any(any(str(cv).lower() in v for v in value_lower) for cv in criterion_values):
                    return False
            case "greater_than":
                try:
                    if criterion.get("has_context", False) and isinstance(value, list):
                        # Context-aware comparison: check each side separately
                        context_criteria_count += 1
                        sides_passed = []
                        
                        for item in value:
                            side = item.get("context", "").lower()
                            try:
                                # Check if this side's value passes the criterion
                                if all(float(item["value"]) > float(cv) for cv in criterion_values):
                                    sides_passed.append(side)
                            except (ValueError, TypeError):
                                # Skip items with non-numeric values
                                continue
                        if affected_side and affected_side.lower() in ["left", "right"]:
                            if affected_side.lower() not in sides_passed:
                                return False
                        else:
                            # No specific side specified - track which sides passed
                            for s in sides_passed:
                                sides_passed_counter[s] += 1
                            if not sides_passed:
                                return False  # Neither side passed
                    else:
                        # Context-agnostic comparison: single value check
                        if any(float(value) <= float(cv) for cv in criterion_values):
                            return False
                except (ValueError, TypeError):
                    return False
                    
            case "less_than":
                try:
                    if criterion.get("has_context", False) and isinstance(value, list):
                        # Context-aware comparison: check each side separately
                        context_criteria_count += 1
                        sides_passed = []
                        
                        for item in value:
                            side = item.get("context", "").lower()
                            try:
                                # Check if this side's value passes the criterion
                                if all(float(item["value"]) < float(cv) for cv in criterion_values):
                                    sides_passed.append(side)
                            except (ValueError, TypeError):
                                continue
                        if affected_side and affected_side.lower() in ["left", "right"]:
                            if affected_side.lower() not in sides_passed:
                                return False
                        else:
                            # No specific side specified - track which sides passed
                            for s in sides_passed:
                                sides_passed_counter[s] += 1
                            if not sides_passed:
                                return False  # Neither side passed
                    else:
                        # Context-agnostic comparison: single value check
                        if any(float(value) >= float(cv) for cv in criterion_values):
                            return False
                except (ValueError, TypeError):
                    return False
            case _:
                continue

    # If context-aware criteria were evaluated, return side info
    if context_criteria_count > 0 and (affected_side is None or affected_side == "" or affected_side.lower() not in ["left", "right"]):
        all_right_passed = "right" in sides_passed_counter and sides_passed_counter["right"] == context_criteria_count
        all_left_passed = "left" in sides_passed_counter and sides_passed_counter["left"] == context_criteria_count
        if all_right_passed and all_left_passed:
            return True, "both"
        elif all_right_passed:
            return True, "right"
        elif all_left_passed:
            return True, "left"
        else:
            return False, None
    elif context_criteria_count > 0 and affected_side and affected_side.lower() in ["left", "right"]:
        return True, affected_side.lower()
    return True

def process_clip_data(condition_clips, processing_criteria, clip_data_columns, all_jsons, file_paths):
    """
    Process JSON files for the given clips.
    Args:
        condition_clips: List of clips in the condition.
        processing_criteria: List of filtering criteria (or single criterion).
        clip_data_columns: List of clip data column dictionaries to populate.
        all_jsons: List to store all downloaded JSON data.
        file_paths: List to store file paths of downloaded JSON data.
    """
    # Ensure processing_criteria is a list
    if not isinstance(processing_criteria, list):
        processing_criteria = [processing_criteria]
    
    for criterion in processing_criteria:
        criterion_source = criterion.get('source', "")
        criterion_processing = criterion.get('processing')
        criterion_label = criterion.get('label', "")
        expected_channel_names = {f"Left{criterion_label}", f"Right{criterion_label}", criterion_label}
        
        # Find columns that match this criterion's source and processing
        matching_columns = [
            col for col in clip_data_columns
            if col.get('processing') == criterion_processing
            and col.get('channel_name') in expected_channel_names
            ]
        
        # Process jsons for each clip
        for clip in condition_clips:
            additional_data = clip.get('additionalData', [])
            file_found = False
            
            for ad in additional_data:
                filename, _ = os.path.splitext(ad['originalFileName'])
                match = re.search(r'<<<(.*?)>>>', filename)
                if not match:
                    continue
                clip_file = match.group(1)
                
                if filename.endswith(criterion_source):
                    file_found = True
                    file_path = f'{clip["projectPath"]}{clip["title"]}/{ad["originalFileName"]}'
                    
                    # First check if we have already downloaded this file
                    if file_path in file_paths:
                        current_index = file_paths.index(file_path)
                        data_channels = all_jsons[current_index]['data']
                    else:
                        file_data = requests_session.get(ad['originalDataDownloadUri']).content
                        readable_data = json.loads(file_data.decode())
                        data_channels = readable_data.get('data', [])
                        all_jsons.append(readable_data)
                        file_paths.append(file_path)
                    
                    if criterion.get("source", "") == "gait_params" and criterion.get("has_context", False):
                        channel_labels = [channel['context'] + channel['label'] for channel in data_channels]
                    else:
                        channel_labels = [channel['label'] for channel in data_channels]
                    
                    # Process only matching columns for this criterion
                    for clip_data_column in matching_columns:
                        label_key = clip_data_column.get('channel_name', "")
                        if label_key in channel_labels:
                            channel_idx = channel_labels.index(label_key)
                            if criterion_source == "gait_params":
                                my_signal = data_channels[channel_idx]['values']["mean"]
                            else:
                                my_signal = [value[clip_file] for value in data_channels[channel_idx]['values']]
                            
                            match criterion_processing:
                                case "start":
                                    value = my_signal[0]
                                case "average":
                                    value = sum(my_signal) / len(my_signal)
                                case "max":
                                    value = max(my_signal)
                                case None:
                                    value = my_signal
                                case "single_stance":
                                    # TODO: implement single stance average
                                    value = DEFAULT_CELL_VALUE
                                case _:
                                    value = DEFAULT_CELL_VALUE
                            clip_data_column["value"].append(value)
                    break  # Found the file for this clip, move to next clip
            
            # If no matching file was found for this clip, append default values
            if not file_found:
                for clip_data_column in matching_columns:
                    clip_data_column["value"].append(DEFAULT_CELL_VALUE)
    
    # Average "value" lists for all columns
    for idx in range(len(clip_data_columns)):
        if len(clip_data_columns[idx]["value"]) > 0:
            # Filter out DEFAULT_CELL_VALUE and convert to numeric
            numeric_values = []
            for val in clip_data_columns[idx]["value"]:
                if val != DEFAULT_CELL_VALUE:
                    try:
                        numeric_values.append(float(val))
                    except (ValueError, TypeError):
                        # Skip non-numeric values
                        pass
            
            # Compute mean only if we have numeric values
            if len(numeric_values) > 0:
                clip_data_columns[idx]["value"] = np.mean(numeric_values)
            else:
                clip_data_columns[idx]["value"] = DEFAULT_CELL_VALUE
        else:
            clip_data_columns[idx]["value"] = DEFAULT_CELL_VALUE
    
    return clip_data_columns

def append_metadata_to_row(params_to_export, data_columns, data, level, source, pre_post: str = "", date_of_birth: str = ""):
    """
    Appends metadata to the data columns for a given level and source.
    Args:
        params_to_export: List of parameters to export.
        data_columns: Dictionary of data columns to append to.
        data: Metadata dictionary.
        level: Level of the data (project, subject, session).
        source: Source of the data (name, subject_metadata, session_metadata).
        pre_post: Suffix to add to column names (e.g., " - Pre", " - Post", or "").
        date_of_birth: Date of birth string in "YYYY-MM-DD" format (used for age calculation).
    """
    for param_to_export in params_to_export:
        if param_to_export.get('level') == level and param_to_export.get('source') == source:
            column_names = []
            if param_to_export.get('has_context'):
                for context in CONTEXT_LABELS:
                    column_names.append(f"{context} {param_to_export.get('column_name')}{pre_post}")
            else:
                column_names.append(f"{param_to_export.get('column_name')}{pre_post}")
            for column_name in column_names:
                match level:
                    case 'project':
                        match source:
                            case 'name':
                                data_columns[column_name] = data['name']
                            case _:
                                continue
                    case 'subject':
                        match source:
                            case 'name':
                                data_columns[column_name] = data['name']
                            case 'group':
                                data_columns[column_name] = data.get('group', DEFAULT_CELL_VALUE)
                            case 'passed_side':
                                data_columns[column_name] = data.get('passed_side', DEFAULT_CELL_VALUE)
                            case 'subject_metadata':
                                data_columns[column_name] = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)
                            case 'intervention_metadata':
                                if param_to_export.get('processing', "") == 'age_at_encounter' and date_of_birth != "":
                                    intervention_date = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)
                                    if intervention_date == DEFAULT_CELL_VALUE:
                                        data_columns[column_name] = DEFAULT_CELL_VALUE
                                        continue
                                    data_columns[column_name] = calculate_age_in_years_and_months(date_of_birth, intervention_date)
                                else:
                                    data_columns[column_name] = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)
                            case 'procedure_metadata':
                                data_columns[column_name] = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)

                            case 'research_projects':
                                data_columns[column_name] = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)
                            case _:
                                continue
                    case 'session':
                        match source:
                            case 'date':
                                if param_to_export.get('processing', "") == 'age_at_encounter' and date_of_birth != "":
                                    data_columns[column_name] = calculate_age_in_years_and_months(date_of_birth, data)
                                else:
                                    data_columns[column_name] = data
                            case 'projectPath':
                                data_columns[column_name] = data
                            case 'session_metadata':
                                metadata_value = data.get(param_to_export.get('label'), DEFAULT_CELL_VALUE)
                                if isinstance(metadata_value, list):
                                    for value in metadata_value:
                                        if column_name.lower().startswith(value.get('context')):
                                            data_columns[column_name] = value.get('value', DEFAULT_CELL_VALUE)
                                elif isinstance(metadata_value, dict) and metadata_value.get("multiselect", False):
                                    data_columns[column_name] = metadata_value.get("value", [])
                                else:
                                    data_columns[column_name] = metadata_value
                    case _:
                        continue
    return data_columns

def append_clip_data_to_row(clip_params, data_columns, data, pre_post: str, all_jsons, file_paths):
    """
    Appends clip/trial data to the data columns for a given clip. 
    Args:
        clip_params: Dictionary of clip-level parameters grouped by source.
        data_columns: Dictionary of data columns to append to.
        data: Clip data.
        pre_post: Suffix to add to column names (e.g., " - Pre", " - Post", or "").
        all_jsons: List to store all downloaded jsons.
        file_paths: List to store file paths of downloaded jsons.
    """
    for source, items in clip_params.items():
        clip_data_columns = []
        for param_to_export in items:
            if param_to_export.get("has_context", False):
                clip_data_columns.append({
                    "channel_name": f'Left{param_to_export.get("label")}',
                    "value": [],
                    "processing": param_to_export.get("processing"),
                    "column_name": f"Left {param_to_export.get('column_name')}{pre_post}"
                })
                clip_data_columns.append({
                    "channel_name": f'Right{param_to_export.get("label")}',
                    "value": [],
                    "processing": param_to_export.get("processing"),
                    "column_name": f"Right {param_to_export.get('column_name')}{pre_post}"
                })
            else:
                clip_data_columns.append({
                    "channel_name": param_to_export.get("label"),
                    "value": [],
                    "processing": param_to_export.get("processing"),
                    "column_name": f"{param_to_export.get('column_name')}{pre_post}"
                })
        # # Process clip-level jsons
        clip_data_columns = process_clip_data(data, items, clip_data_columns, all_jsons, file_paths)
        # Add the value in clip_data_columns to data_columns
        for clip_data_column in clip_data_columns:
            column_name = clip_data_column.get("column_name", "")
            if column_name in data_columns:
                data_columns[column_name] = clip_data_column.get("value", DEFAULT_CELL_VALUE)
        
    return data_columns

def process_subjects(subjects, api):
    """
    Process a list of subject dictionaries by applying filtering and transformation steps.
    Subjects are filtered based on inclusion criteria at potentially three different levels:
    1. Subject-level filtering based on subject metadata.
    2. Session-level filtering based on session metadata and/or pre/post intervention criteria.
    3. Clip-level filtering based on specific session characteristics.
    Args:
        subjects: List of subject dictionaries to process.
        api: An instance of the API client to fetch additional data if needed.
    Returns a list of filtered patient dictionaries with their sessions.
    """

    grouped_inclusion_criteria = group_inclusion_criteria_by_source(INCLUSION_CRITERIA)

    if len(grouped_inclusion_criteria.get("subject_metadata", [])) == 0:
        # If no subject-level inclusion criteria, skip first layer of filtering
        filtered_subjects_step_1 = subjects
    else:
        ## First layer of filtering based on inclusion criteria (subject level)
        filtered_subjects_step_1 = []
        for subject in subjects:      
            subject_metadata = {}
            if 'metadata' in subject and subject['metadata'] is not None: # Also checking that metadata is not None
                subject_metadata = json.loads(subject['metadata'])  
            # First layer of filtering based on inclusion criteria (subject level)
            if not filter_based_on_inclusion_criteria(subject_metadata, INCLUSION_CRITERIA, "subject_metadata"):
                continue
            
            filtered_subjects_step_1.append(subject)

        logging.info('Number of subjects after first layer (subject metadata) of filtering: %d', len(filtered_subjects_step_1))
    
    
    ## Second layer of filtering based on inclusion criteria (session and interventionlevel)
    filtered_subjects_step_2 = []
    for subject in filtered_subjects_step_1:   
        subject_metadata = {}
        if 'metadata' in subject and subject['metadata'] is not None:
            subject_metadata = json.loads(subject['metadata'])
        
        sessions = sorted(subject["sessions"], key=lambda x: x["date"])
        if len(sessions) < 2:
            # we need at least 2 sessions to have pre and post intervention data
            continue
        process_consecutive_session_pairs(
            subject,
            subject_metadata,
            sessions,
            INCLUSION_CRITERIA,
            INTERVENTION_CRITERIA,
            MIN_MONTHS_BETWEEN_SESSIONS,
            MAX_YEARS_AFTER_INTERVENTION,
            filtered_subjects_step_2
        )
        
    logging.info('Number of session-pairs after second layer of filtering: %d', len(filtered_subjects_step_2))

    ## Third layer of filtering based on inclusion criteria (clip level)
    # Now that the dataset has been narrowed down based on subject and session level criteria,
    # we can retrieve clip-level data and apply clip-level filtering criteria. This is done after
    # the previous filtering steps to minimize the number of API calls and amount of data processing
    # needed at the clip level, which can be resource intensive.
    
    # Extract unique subjects (from subject['id']) from filtered_subjects_step_2
    unique_subject_ids = []
    for subject in filtered_subjects_step_2:
        if subject['id'] not in unique_subject_ids:
            unique_subject_ids.append(subject['id'])

    # First retrieve subjects with clips containing json data
    # Parallel retrieval of subject data
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        subjects_with_clips = list(executor.map(api.getSubjectData, unique_subject_ids))

    # Get required clip data sources dynamically
    clip_data_sources = get_clip_data_sources(PARAMS_TO_EXPORT, INCLUSION_CRITERIA) 
    URLs = []
    file_paths = []
    for subject_details in subjects_with_clips:
        for session in subject_details.get("sessions", []):
            for c in session.get("clips", []):
                project_path = c.get("projectPath", "")
                condition_name = os.path.basename(project_path.rstrip('/'))
                if CONDITION_TARGET_NAMES and condition_name.lower() not in CONDITION_TARGET_NAMES:
                    continue
                for ad in c.get("additionalData", []):
                    if any(ad["originalDataDownloadUri"].endswith(f"{source}.json") for source in clip_data_sources):
                        URLs.append(ad["originalDataDownloadUri"])
                        file_paths.append(f'{c["projectPath"]}{c["title"]}/{ad["originalFileName"]}')

    # Download all JSON files in parallel
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        all_jsons = list(executor.map(download_with_session, URLs))
    
    # repeat session filtering (step 2) to get session pairs with clips after retrieving clip data
    filtered_subjects_step_2 = []
    for subject in subjects_with_clips:   
        subject_metadata = {}
        if 'metadata' in subject and subject['metadata'] is not None:
            subject_metadata = json.loads(subject['metadata'])
        
        sessions = sorted(subject["sessions"], key=lambda x: x["date"])
        if len(sessions) < 2:
            continue
        process_consecutive_session_pairs(
            subject,
            subject_metadata,
            sessions,
            INCLUSION_CRITERIA,
            INTERVENTION_CRITERIA,
            MIN_MONTHS_BETWEEN_SESSIONS,
            MAX_YEARS_AFTER_INTERVENTION,
            filtered_subjects_step_2
        )
    
    clip_inclusion_criteria = group_clip_params_by_source(INCLUSION_CRITERIA)
    if len(clip_inclusion_criteria) == 0:
        # If no clip-level inclusion criteria, skip third layer of filtering
        filtered_subjects_step_3 = filtered_subjects_step_2
    else:
        filtered_subjects_step_3 = []
        for subject in filtered_subjects_step_2:
            valid_sessions = [] # To store sessions that pass the clip-level filtering
            n_added_sessions = 0
            passed_sides_list = []
            for session in subject["sessions"]:
                session_id = session['id']
                clips = session.get('clips', [])
                if not clips or len(clips) == 0:
                    clips = api.getSessionClips(session_id)
                    session['clips'] = clips
                has_valid_clips = True
                condition_clips = []
                for clip in clips:
                    project_path = clip.get("projectPath", "")
                    condition_name = os.path.basename(project_path.rstrip('/'))
                    if not CONDITION_TARGET_NAMES or condition_name.lower() in CONDITION_TARGET_NAMES:
                        condition_clips.append(clip)
                
                clip_data_columns = []
                for criterion in INCLUSION_CRITERIA:
                    if criterion.get("level", "") == "clip":
                        if criterion.get("has_context", False):
                            clip_data_columns.append({
                                "channel_name": f'Left{criterion.get("label")}',
                                "value": [],
                                "processing": criterion.get("processing"),
                            })
                            clip_data_columns.append({
                                "channel_name": f'Right{criterion.get("label")}',
                                "value": [],
                                "processing": criterion.get("processing")
                            })
                        else:
                            clip_data_columns.append({
                                "channel_name": criterion.get("label"),
                                "value": [],
                                "processing": criterion.get("processing")
                            })
                        # # Process clip-level jsons
                        clip_data_columns = process_clip_data(condition_clips, criterion, clip_data_columns, all_jsons, file_paths)
                        
                        # last layer of filtering based on trial level inclusion criteria
                        result = filter_based_on_inclusion_criteria(clip_data_columns, INCLUSION_CRITERIA, criterion.get('source', ""), n_added_sessions, subject.get("passed_side", subject_metadata.get("subject-diagnosis-laterality", None)))
                        if isinstance(result, tuple):
                            passed_filter, passed_sides = result
                            passed_sides_list.append(passed_sides)
                        else:
                            passed_filter = result
                            passed_sides = None
                            
                        if not passed_filter:
                            has_valid_clips = False
                            break
                
                if not has_valid_clips:
                    break

                if has_valid_clips and len(condition_clips) > 0:    
                    valid_sessions.append(session)
                    n_added_sessions += 1
                    
            if has_valid_clips and n_added_sessions == 2 and len(condition_clips) > 0:
                subject["sessions"] = valid_sessions
                # Determine passed_sides for the subject. If both sides have passed criteria, select 'right' by default
                unique_sides = set(passed_sides_list)
                if unique_sides <= {"both", "right"}:
                    final_side = "right"
                elif unique_sides <= {"both", "left"}:
                    final_side = "left"
                elif "left" in unique_sides and "right" in unique_sides:
                    # Skip this subject due to mixed sides
                    continue

                subject["passed_side"] = final_side
                filtered_subjects_step_3.append(subject)
                
        logging.info('Number of consecutive session pairs after third layer (i.e., clip-level) of filtering: %d', len(filtered_subjects_step_3))  


    ## Extract data columns for the final filtered dataset
    data_columns = define_column_names(PARAMS_TO_EXPORT)
    intervention_data_columns = define_column_names(INTERVENTION_PARAMS_TO_EXPORT)
    
    logging.info('Extracting output data')
    for subject in filtered_subjects_step_3:
        row = {col: None for col in data_columns.keys()}
        subject_metadata = {}
        if 'metadata' in subject and subject['metadata'] is not None:
            subject_metadata = json.loads(subject['metadata'])
        
        # extract research project data from subject metadata
        research_projects = subject_metadata.get("Research projects", [])
        # extract intervention data for intervention data sheet
        interventions = subject_metadata.get("interventions", []) # List of interventions
        for intervention in interventions:
            row_intervention = {col: None for col in intervention_data_columns.keys()}
            procedures = intervention.get("procedures", [])
            for procedure in procedures:
                if len(research_projects) > 0:
                    # currently we only consider the first research project. TODO: handle multiple research projects 
                    row_intervention = append_metadata_to_row(INTERVENTION_PARAMS_TO_EXPORT, row_intervention, research_projects[0], 'subject', 'research_projects',  "")
                row_intervention = append_metadata_to_row(INTERVENTION_PARAMS_TO_EXPORT, row_intervention, subject, 'subject', 'name', "")
                row_intervention = append_metadata_to_row(INTERVENTION_PARAMS_TO_EXPORT, row_intervention, intervention, 'subject', 'intervention_metadata', date_of_birth= subject_metadata.get("subject-date-of-birth", ""))
                row_intervention = append_metadata_to_row(INTERVENTION_PARAMS_TO_EXPORT, row_intervention, procedure, 'subject', 'procedure_metadata', "")
                for col in intervention_data_columns.keys():
                    intervention_data_columns[col].append(row_intervention.get(col, DEFAULT_CELL_VALUE))

        row = append_metadata_to_row(PARAMS_TO_EXPORT, row, subject, 'subject', 'name', "")
        row = append_metadata_to_row(PARAMS_TO_EXPORT, row, subject_metadata, 'subject', 'subject_metadata', "")
        row = append_metadata_to_row(PARAMS_TO_EXPORT, row, subject, 'subject', 'group', "")
        row = append_metadata_to_row(PARAMS_TO_EXPORT, row, subject, 'subject', 'passed_side', "")
        n_added_sessions = 0
        for session in subject["sessions"]:
            session_id = session['id']
            
            session_date = (
                datetime.fromisoformat(session["date"]).strftime("%Y-%m-%d")
                if session.get("date")
                else DEFAULT_CELL_VALUE
            )
            my_session_metadata = {}
            if 'metadata' in session and session['metadata'] is not None:
                session_metadata = json.loads(session['metadata'])
                if 'metadata' in session_metadata.keys() and session_metadata['metadata'] is not None:
                    my_session_metadata = session_metadata['metadata']
   
                if n_added_sessions == 0:
                    pre_post = " - Pre"
                    row = append_metadata_to_row(PARAMS_TO_EXPORT, row, session_date, 'session', 'date', pre_post, subject_metadata.get("subject-date-of-birth", ""))
                    row = append_metadata_to_row(PARAMS_TO_EXPORT, row, my_session_metadata, 'session', 'session_metadata', pre_post)

                elif n_added_sessions == 1:
                    pre_post = " - Post"
                    row = append_metadata_to_row(PARAMS_TO_EXPORT, row, session_date, 'session', 'date', pre_post, subject_metadata.get("subject-date-of-birth", ""))
                    row = append_metadata_to_row(PARAMS_TO_EXPORT, row, my_session_metadata, 'session', 'session_metadata', pre_post)

                else:
                    logging.warning('More than two sessions found for subject %s. Only the first two sessions are considered.', subject.get('name', 'Unknown'))
                
                clips = session.get('clips', [])
                condition_clips = []
                for clip in clips:
                    project_path = clip.get("projectPath", "")
                    condition_name = os.path.basename(project_path.rstrip('/'))
                    if not CONDITION_TARGET_NAMES or condition_name.lower() in CONDITION_TARGET_NAMES:
                        condition_clips.append(clip)
                
                
                if n_added_sessions == 0:
                    pre_post = " - Pre"
                elif n_added_sessions == 1:
                    pre_post = " - Post"
                else:
                    break
                
                clip_params = group_clip_params_by_source(PARAMS_TO_EXPORT)
                row = append_clip_data_to_row(clip_params, row, condition_clips, pre_post, all_jsons, file_paths)
                n_added_sessions += 1
                        
            
        if n_added_sessions == 2:
            for col in data_columns.keys():
                data_columns[col].append(row.get(col, DEFAULT_CELL_VALUE))

    return data_columns, intervention_data_columns

def process_consecutive_session_pairs(
    subject,
    subject_metadata,
    sessions,
    INCLUSION_CRITERIA,
    INTERVENTION_CRITERIA,
    MIN_MONTHS_BETWEEN_SESSIONS,
    MAX_YEARS_AFTER_INTERVENTION,
    filtered_subjects_step_2
):
    """
    Helper function to process all consecutive session pairs for a subject.
    It checks whether each pair meets the specified criteria regarding time between sessions, interventions, and session metadata.
    If a valid pair is found, it creates a new subject entry with the filtered sessions and metadata.
    Appends valid subjects with filtered sessions to filtered_subjects_step_2.
    """
    for i in range(len(sessions) - 1):
        filtered_sessions = []
        session_pair = [sessions[i], sessions[i + 1]]
        valid_pair = True

        # Get session dates and make sure they are within the date range
        try:
            date1_dt = datetime.fromisoformat(session_pair[0]["date"]).replace(tzinfo=None)
            if (STARTING_DATE_DT and date1_dt < STARTING_DATE_DT) or (ENDING_DATE_DT and date1_dt > ENDING_DATE_DT):
                continue  # Skip this pair
            date2_dt = datetime.fromisoformat(session_pair[1]["date"]).replace(tzinfo=None)
            if (STARTING_DATE_DT and date2_dt < STARTING_DATE_DT) or (ENDING_DATE_DT and date2_dt > ENDING_DATE_DT):
                continue  # Skip this pair
        except Exception:
            continue
        
        # Check if sessions meet minimum time separation requirement
        if MIN_MONTHS_BETWEEN_SESSIONS is not None and MIN_MONTHS_BETWEEN_SESSIONS > 0:
            if date2_dt < date1_dt + relativedelta(months=MIN_MONTHS_BETWEEN_SESSIONS):
                continue  # Sessions too close together, skip this pair

        # Find and evaluate interventions between session dates
        interventions = subject_metadata.get("interventions", [])
        is_control = True
        intervention_groups_found = set()
        exclusion_found = False

        for intervention in sorted(interventions, key=lambda x: x.get("date", "")):
            intervention_date_str = intervention.get("date", "")
            if not intervention_date_str or not date1_dt or not date2_dt:
                continue
            intervention_date_dt = datetime.fromisoformat(intervention_date_str).replace(tzinfo=None)
            if date1_dt < intervention_date_dt < date2_dt:
                # Check if post-session is within acceptable timeframe after intervention
                if MAX_YEARS_AFTER_INTERVENTION is not None and MAX_YEARS_AFTER_INTERVENTION > 0:
                    if date2_dt > intervention_date_dt + relativedelta(years=MAX_YEARS_AFTER_INTERVENTION):
                        continue  # Post-session too long after intervention, skip this pair

                is_control = False
                procedures = intervention.get("procedures", [])

                # 1. Check exclusion criteria first
                for exclusion_set in INTERVENTION_CRITERIA["exclusion"]:
                    exclusion_match = True
                    for criterion in exclusion_set:
                        label = criterion["label"]
                        values = criterion["value"] if isinstance(criterion["value"], list) else [criterion["value"]]
                        operation = criterion["operation"]
                        criterion_met = False
                        for procedure in procedures:
                            proc_value = procedure.get(label, "")
                            if operation == "is":
                                if str(proc_value).lower() in [str(v).lower() for v in values]:
                                    criterion_met = True
                                    break
                            elif operation == "is_not":
                                if str(proc_value).lower() not in [str(v).lower() for v in values]:
                                    criterion_met = True
                                    break
                        if not criterion_met:
                            exclusion_match = False
                            break  # This exclusion_set is not fully met, so don't skip
                    if exclusion_match:
                        exclusion_found = True
                        break  # All criteria in this exclusion_set are met, skip this session pair

                # 2. Check inclusion criteria
                for inclusion_set in INTERVENTION_CRITERIA["inclusion"]:
                    inclusion_match = True
                    group_name = None
                    for criterion in inclusion_set:
                        label = criterion["label"]
                        values = criterion["value"] if isinstance(criterion["value"], list) else [criterion["value"]]
                        operation = criterion["operation"]
                        group_name = criterion.get("group", None)
                        criterion_met = False
                        for procedure in procedures:
                            proc_value = procedure.get(label, "")
                            if operation == "is":
                                if str(proc_value).lower() in [str(v).lower() for v in values]:
                                    criterion_met = True
                                    break
                            # You can add more operations here if needed
                        if not criterion_met:
                            inclusion_match = False
                            break  # If any criterion is not met, this inclusion set does not match
                    if inclusion_match and group_name:
                        intervention_groups_found.add(group_name)

        if exclusion_found or (not is_control and len(intervention_groups_found) == 0):
            continue  # Skip this session pair

        group_str = ", ".join(sorted(intervention_groups_found)) if intervention_groups_found else "Control"
        # If intervention_found is False, treat as control (no intervention between sessions)

        # Filter sessions based on session_metadata criteria
        # Track which sides passed for each session to ensure consistency
        passed_sides_list = []
        for n_added_sessions, session in enumerate(session_pair):
            # Extract session metadata
            my_session_metadata = {}
            if 'metadata' in session and session['metadata'] is not None:
                session_metadata = json.loads(session['metadata'])
                if 'metadata' in session_metadata.keys() and session_metadata['metadata'] is not None:
                    my_session_metadata = session_metadata['metadata']
            result = filter_based_on_inclusion_criteria(
                my_session_metadata,
                INCLUSION_CRITERIA,
                "session_metadata",
                n_added_sessions,
                subject_metadata.get("subject-diagnosis-laterality", None)
            )
            
            if isinstance(result, tuple):
                passed_filter, passed_sides = result
                passed_sides_list.append(passed_sides)
            else:
                passed_filter = result
                passed_sides = None
                
            if not passed_filter:
                valid_pair = False
                break
                
            filtered_sessions.append(session)
            
        if valid_pair and len(filtered_sessions) == 2:
            # Determine passed_sides for the subject
            unique_sides = set(passed_sides_list)
            if unique_sides == {"both"}:
                final_side = "both"
            elif unique_sides <= {"both", "right"}:
                final_side = "right"
            elif unique_sides <= {"both", "left"}:
                final_side = "left"
            elif "left" in unique_sides and "right" in unique_sides:
                # Skip this subject due to mixed sides
                continue
            
            # Create a new subject entry with the filtered session pair
            new_subject = copy.deepcopy(subject)
            new_subject["sessions"] = filtered_sessions
            new_subject['group'] = group_str
            new_subject['passed_side'] = final_side
            filtered_subjects_step_2.append(new_subject)


## ============================================================================
## EXCEL EXPORT AND FORMATTING
## ============================================================================

def save_to_excel(df: pd.DataFrame, filename: str, sheet_name: str = 'Data') -> None:
    """
    Save DataFrame to Excel with professional formatting.
    
    Applies:
    - Colored header row with white text
    - Auto-sized columns (with max width limit)
    - Frozen top row for scrolling
    - Table format with filtering enabled
    
    Args:
        df: DataFrame to export
        filename: Output file path
        sheet_name: Name of the Excel sheet (default: 'Data')
    """
    if df.empty:
        logging.warning(f"DataFrame is empty, no data to export to '{filename}'")
        return
    
    # Write DataFrame to Excel
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Apply formatting
    _apply_excel_formatting(filename, sheet_name, df)
    
    logging.info(f"Excel file saved successfully: {filename}")


def _apply_excel_formatting(filename: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Apply formatting to an existing Excel file.
    
    Args:
        filename: Path to Excel file
        sheet_name: Name of sheet to format
        df: DataFrame used to determine dimensions
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    
    # Format header row
    _format_header_row(ws)
    
    # Auto-size columns
    _auto_size_columns(ws, df)
    
    # Add filter table
    _add_filter_table(ws, df, sheet_name)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filename)


def _format_header_row(worksheet) -> None:
    """
    Apply formatting to the header row (row 1).
    
    Args:
        worksheet: Openpyxl worksheet object
    """
    header_fill = PatternFill(
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
        fill_type='solid'
    )
    header_font = Font(bold=True, color=HEADER_FONT_COLOR)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _auto_size_columns(worksheet, df: pd.DataFrame) -> None:
    """
    Automatically size columns based on content width.
    
    Args:
        worksheet: Openpyxl worksheet object
        df: DataFrame used to calculate column widths
    """
    for col_num, col_header in enumerate(df.columns, 1):
        # Start with header width
        max_length = len(str(col_header))
        
        # Check data rows (limit check to first 100 rows for performance)
        max_row = min(worksheet.max_row, MAX_ROWS_FOR_WIDTH_CALC + 1)
        for row_num in range(2, max_row + 1):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width with padding and max limit
        col_letter = get_column_letter(col_num)
        adjusted_width = min(max_length + COLUMN_PADDING, MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def _add_filter_table(worksheet, df: pd.DataFrame, sheet_name: str) -> None:
    """
    Add Excel table with filters.
    
    Args:
        worksheet: Openpyxl worksheet object
        df: DataFrame used to determine table range
        sheet_name: Sheet name (used for table name)
    """
    if df.empty:
        return
    
    # Define table range (A1 to last column/row)
    last_col = get_column_letter(df.shape[1])
    last_row = df.shape[0] + 1
    table_ref = f'A1:{last_col}{last_row}'
    
    # Create and add table
    table_name = sheet_name.replace(' ', '_').replace('-', '_')
    table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_ref)
    worksheet.add_table(table)


## ============================================================================
## MAIN EXECUTION
## ============================================================================

def main():  
    """
    Main execution function that:
    1. Initializes API connection
    2. Queries and processes subjects from specified projects
    3. Exports results to Excel files
    """
    ## Setup the API
    # Load config
    personal_config = os.path.join(parent_folder, "mvshlf-config.json")
    if not os.path.isfile(personal_config):
        raise FileNotFoundError(
            f"Configuration file '{personal_config}' is missing.\n"
            "Ensure the file exists with the correct name and path."
        )

    with open(personal_config, "r") as config_file:
        data = json.load(config_file)

    custom_timeout = 600 # timeout (integer) for HTTP calls. Defaults to 120 if undefined

    # Initialize the API with custom timeout
    api = MoveshelfApiCustomized(
        api_key_file=os.path.join(parent_folder, data["apiKeyFileName"]),
        api_url=data["apiUrl"],
        timeout=custom_timeout,
    )

    # Increase connection pool size globally for urllib3 to match MAX_WORKERS thread workers
    api.http.connection_pool_kw['maxsize'] = POOL_MAXSIZE

    projects = api.getUserProjects()    
    project_names = [project['name'] for project in projects]

    for pname in PROJECT_NAMES:
        if pname not in project_names:
            logging.error('Could not find project in user projects list: %s', pname)
            logging.info('Available projects: %s', ', '.join(project_names))
            return

    # Initialize combined data columns for all projects
    combined_data_columns = {}
    combined_intervention_data_columns = {}
    
    if query_projects_in_parallel:
        list_of_project_ids = []
        for pname in PROJECT_NAMES:
            i_project = project_names.index(pname)
            my_project_id = projects[i_project]['id']
            list_of_project_ids.append(my_project_id)

        # Query all subjects in parallel
        logging.info('Extracting filtered subjects (without data files), optionally filtered by session date, from projects in parallel: %s', PROJECT_NAMES)
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            subjects_in_parallel = list(executor.map(api.getFilteredProjectSubjects, list_of_project_ids, [subject_metadata_filters]*len(list_of_project_ids), [session_filters]*len(list_of_project_ids), [False]*len(list_of_project_ids)))
        logging.info('Total number of subjects retrieved from all projects: %d', sum(len(project_subjects) for project_subjects in subjects_in_parallel))

        # Process subjects in parallel using ThreadPoolExecutor
        logging.info('Processing subjects in parallel to extract session-pairs...')
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            results = list(executor.map(process_subjects, subjects_in_parallel, [api]*len(subjects_in_parallel)))
        logging.info('Data extraction completed for all subjects. Combining results and exporting to Excel.')

        # results is a list of tuples: [(data_columns1, intervention_data_columns1), (data_columns2, intervention_data_columns2), ...]
        # Combine all data_columns and intervention_data_columns
        for data_cols, intervention_cols in results:
            for k, v in data_cols.items():
                combined_data_columns.setdefault(k, []).extend(v)
            for k, v in intervention_cols.items():
                combined_intervention_data_columns.setdefault(k, []).extend(v)
    else:
        # Query and process projects sequentially. Combine results from all projects before conversion to excel
        for pname in PROJECT_NAMES:
            i_project = project_names.index(pname)
            my_project_id = projects[i_project]['id']

            logging.info('Extracting filtered subjects (without data files), optionally filtered by session date, from project: %s', pname)
            subjects = api.getFilteredProjectSubjects(my_project_id, subject_metadata_filters=subject_metadata_filters, session_filters=session_filters, include_additional_data=False)
            logging.info('Total number of subjects in project within the date range: %d', len(subjects))
            
            logging.info('Processing subjects to extract session-pairs for project: %s', pname)
            data_columns, intervention_data_columns = process_subjects(subjects, api)
            logging.info('Data processing completed for project: %s', pname)
            
            # Combine results from this project
            for k, v in data_columns.items():
                combined_data_columns.setdefault(k, []).extend(v)
            for k, v in intervention_data_columns.items():
                combined_intervention_data_columns.setdefault(k, []).extend(v)

    ## Export data
    df_data = pd.DataFrame(combined_data_columns)
    save_to_excel(df_data, DATA_SPREADSHEET_FILENAME, sheet_name='Data')
    df_intervention = pd.DataFrame(combined_intervention_data_columns)
    save_to_excel(df_intervention, INTERVENTION_DATA_SPREADSHEET_FILENAME, sheet_name='Intervention Data')

if __name__ == '__main__':
    main()